import requests
from time import sleep
from bs4 import BeautifulSoup as bs
import re
import os
from openpyxl import load_workbook
from selenium import webdriver
from podcast_dict import podcasts_dict
import urllib3
import io
import xml.dom.minidom
from alma_tools_v3 import AlmaTools

urllib3.disable_warnings()
# driver = webdriver.Firefox()
##################################################################################################


def main():

	"""Example of usage"""

	mms_id = ""

	my_api = AlmaTools("sb")
	# my_api.get_bib("9919000573502836")
	# print(my_api.xml_response_data)
	# item_pid = "23376150750002836"
	# po_line = "245690-ilsdb"
	# my_api.get_items_by_po_line( po_line, options={})

	# # mis_mms_list = ["999999999999"]

	# # for mms in mis_mms_list:
	# # 		my_api.delete_bib(mms)
	# print(my_api.xml_response_data.encode("utf-8"))
	#######################################GET BIB RECORD#######################################
	# my_api.get_bib("9919173839702836")#, {"limit":"100"})
	# print(my_api.xml_response_data)
	#######################################INVOICES################################################
	#my_api.get_invoices()
	# invoice_id = "11882474040002836"
	# invoice_id = "11882474040002836"
	# invoice_id = "10581159270002836"
	#invoice_id = "11882474040002836"
	# my_api.get_invoice(invoice_id)
	# # new_data= my_api.xml_response_data
	# my_api.process_invoice(invoice_id)
	# print(my_api.xml_response_data)
	# new_data = my_api.xml_response_data.replace('desc="In-Review"','desc="Ready"').replace(">INREVIEW<",">READY<").replace('desc="Pending"','desc="Approved"').replace(">PENDING<",">APPROVED<").replace('<invoice_workflow_status desc="In Review">InReview</invoice_workflow_status>','<invoice_workflow_status/>')#..replace('desc="Active"','desc="Closed"').replace(">ACTIVE<",">CLOSED<")
	# new_data = new_data.replace('</invoice_approval_status><additional_charges>','</invoice_approval_status><approved_by>Cousens Dale*</approved_by><approval_date>2021-09-09Z</approval_date><additional_charges>')
	# print(new_data)
	# my_api.update_invoice(invoice_id, new_data)
	# print(my_api.xml_response_data)
	# my_api.get_invoice_lines(invoice_id)
	# #print(my_api.xml_response_data)
	# lines = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)

	# for line in lines:
	# 	print(line)
	# 	my_api.get_invoice_line(invoice_id, line)
	# 	#print(my_api.xml_response_data)
		
	# 	new_data = my_api.xml_response_data.replace("<percent>100.00</percent>","<percent></percent>")#.replace('desc="In-Review"','desc="Ready"').replace(">INREVIEW<",">READY<")		print(new_data)
	# 	my_api.update_invoice_line(invoice_id, line, new_data)
	# 	print(my_api.xml_response_data)
	# 	quit()
	# invoice_template = r"Y:\ndha\pre-deposit_prod\LD_working\invoice_test\invoice_line_template_short.xml"
	# with open (invoice_template, "r") as f:
	# 	data = f.read()
	# my_api.create_invoice_line("11882474040002836",data, {'op':'process_invoice'}
	# print(my_api.xml_response_data)

	#####################################
	# print(my_api.xml_response_data.encode("utf-8"))
	# dom = xml.dom.minidom.parse(io.StringIO(my_api.xml_response_data)) # or xml.dom.minidom.parseString(xml_string)
	# pretty_xml_as_string = dom.toprettyxml()
	# print(pretty_xml_as_string)
	# #######################################
	# mms_id = "9919119906902836"
	# my_hold = "22369219170002836"
	# item = "23380731890002836"

	# my_api.get_item(mms_id, my_hold,item)
	# print(my_api.xml_response_data)
	# #######################################
	# my_api.update_bib(mms_id, my_api.xml_record_data)
	# print(my_api.status_code)
	#######################################
	# my_api.get_representations("9915184833502836",{"limit":'100',"offset":"500"})
	# print(my_api.xml_response_data.encode("utf-8"))
	###############################################
	# my_api.get_portfolios("9919000573502836")#,"53354322810002836")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	# ###########################################PORTFOLIOS#########################################

	# my_api.get_portfolio("9918122880502836","53376296880002836")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	#########################################
	# my_api.get_ecollection("9918748064302836","61325625670002836")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	# ####################################################################
	# my_api.get_po_line("POL-180272")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	# my_api.get_item("9918137053802836","22278070670002836","23319192270002836")
	# print(my_api.xml_response_data)

	# my_api.receive_item("POL-76418","23319192270002836",my_api.xml_response_data,{"op":"receive"})
	# print(my_api.xml_response_data)
	# my_api.update_po_line("POL-76418",my_api.xml_response_data)
	# print(my_api.xml_response_data)
	####################################################################
	# my_api.get_item("9918166772702836","22294172070002836","23361571080002836")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	# ####################################################################
	# my_api.get_po_line("POL-158984")
	# print(my_api.xml_response_data)
	# print(my_api.status_code)
	#######################################################################
	# with open("items_data_test.txt","r") as f:
	# 	data=f.read()
	# print(data)
	# my_api.create_item("9914758883502836","22254819960002836",data, {"generate_description":True})
	# print(my_api.xml_response_data)
	
	####################################################################
	# with open(r"Y:\ndha\pre-deposit_prod\LD_working\issuu_main\assets\templates\holding.xml","r") as f:
	# 	data = f.read()
	# my_api.create_holding("",data)
	# print(my_api.xml_response_data)

	####################################################################
	# my_list = [["9918166769702836","22294170890002836","23361559820002836"],["9918166769702836","22294170890002836","23361570070002836"],["9918166772702836","22294172070002836","23361571080002836"],["9918166769902836","22294170920002836","23361571070002836"],["9918166769502836","22294170850002836","23361559680002836"],["9918963672702836","22346162470002836","23361559670002836"],["9918166769602836","22294170870002836","23361571060002836"],["9916482513502836","22218301930002836","23361559620002836"],["9918166770602836","22294170970002836","23361557880002836"],["9918166772902836","22294172110002836","23361559600002836"],["9913788543502836","22194217920002836","23361571050002836"]]
	# for lst in my_list:
	# 	mms_id= lst[0]
	# 	holding_id = lst[1]
	# 	item_pid = lst[2]
	# 	my_api.delete_item(mms_id, holding_id, item_pid)
	# 	print(my_api.xml_response_data)
	##################################################################################
	# file_path = r"mms_to_delete.txt"
	# with open(file_path,"r") as f:
	# 	data = f.read()
	# for el in data.split("\n"):
	# 	my_api.delete_bib(el)
	#################################################################################
	# mms_list = ["9919077172602836" ,"9919078573602836","9919078670802836"]
	# for mms in mms_list:
	# 	my_api.delete_bib(mms)
	# 	print(my_api.xml_response_data)

	###############################FROM ALMA SET FULL VIEW SPREADSHEET#########################################################
	# workook_path = r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\assets\results_Kelli.xlsx"
	# if os.path.exists(workook_path):


	# 	wb = load_workbook(workook_path)
	# 	#Enter name of the working sheet below
	# 	ws= wb.get_sheet_by_name("results")
	# 	#if now headers min_row =1
	# 	for row in ws.iter_rows(min_row=2):
	# 	#21for full results
	# 	#depending on where mms id is row[3] should be changed to number of column started from 0.
	# 		mms = row[26].value
	# 		my_api.delete_bib(row[26].value)
	# 		print(my_api.xml_response_data)
	###############################################GET REPRESENTATIONS#################3
	# my_dict = {}
	# rosetta_report_path = r"D:ddd_query.xlsx"
	# if os.path.exists(rosetta_report_path):
	# 	wb = load_workbook(rosetta_report_path)
	# 	#Enter name of the working sheet below
	# 	ws= wb.get_sheet_by_name("ddd_query")
	# 	#if now headers min_row =1
	# 	for row in ws.iter_rows(min_row=2):
	# 		mms = row[2].value
	# 		ie = row[1].value
	# 		label = row[10].value

	# 		print(mms)
	# 		print(label)
	# 		print(ie)
	# 		my_label = "Session "+label.split("-")[-1].lstrip(" Session").split(" ")[0]
	# 		print(my_label)
	# 		my_dict[ie] =my_label
	# for mms in os.listdir(r"Y:\ndha\pre-deposit_prod\server_side_deposits\prod\ld_scheduled\oneoff_audio\9919014268602836\content"):
	# 	my_api.get_representations(mms.split(".")[0])
	# 	print(mms.split(".")[0])
	# # 	print(my_api.xml_response_data)\


	# my_api.get_representations(mms_id,{"limit":"100"})
	# # print(my_api.xml_response_data)
	# num = re.findall(r'record_count="(.*?)">',my_api.xml_response_data)[0]
	# # print(num)
	# for i in range((int(num)//100)+1):
		
	# 	my_api.get_representations(mms_id,{"limit":"100","offset":str(100*i)})
	# 	repres = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)
	# 	for rep in repres:
	# 		my_api.get_representation (mms_id,rep )
	# 		#print(my_api.xml_response_data)
	# 		ie = re.findall(r"pubam:(.*?)</",my_api.xml_response_data)[0]
	# 		label = re.findall(r'<label>(.*?)</label', my_api.xml_response_data)[0]
	# 		delivery_url = re.findall(r'<delivery_url>(.*?)</delivery_url>', my_api.xml_response_data)[0]
	# 		if "2017" in label:
			# 	#new_label = re.findall(r'<label>(.*?)</label', my_api.xml_response_data)[0].replace(")","").replace("(","").replace(". ",".")
				#my_data = my_api.xml_response_data.replace("2022","2021")#.replace("2022 <",">2021<")
				#my_api.update_representation(mms_id, rep, my_data)
				#print(my_api.xml_response_data)
				# print(ie)
				# print(label)
				# print(delivery_url)

	########################################GET REPRESENTATION####################################
	# my_api.get_representation ("9918991865302836", "32363085860002836")
	# print(my_api.xml_response_data)
	########################################GET PO_Lines by mms##############################################
	#my_api.get_po_lines({"q":"mms_id~9919124598602836","status":"ACTIVE","limit":"100","offset":"0","order_by":"title","direction":"desc","expand":"LOCATIONS"})
	#print(my_api.xml_response_data)
	######################################## GET SET MEMBERS##############################

	# my_api.get_set_members("10799432720002836",{"limit":"100"})
	# print(my_api.xml_response_data)
	# bibs = re.findall(r"<id>(.*?)</id>", my_api.xml_response_data)	
	# print(bibs)
	# titles = re.findall(r'<description>(.*?)</description>', my_api.xml_response_data)
	# print(len(bibs))
	# for ind in range(len(bibs)):
	# 	print(bibs[ind])
	# 	print(titles[ind])

#######################################GET REPRESENTATIONS####################################
	# mms_id = "9918755167602836"

	# my_api.get_representations(mms_id,{"limit":"100"})
	# # print(my_api.xml_response_data)
	# labels = {}
	# labels_to_delete = {}
	# try:
	# 	total_count = re.findall(r'count="(.*?)">',my_api.xml_response_data)[0]
	# 	# print(total_count)
	# 	print(mms_id)
	# except Exception as e:
	# 	# print(my_api.xml_response_data)
	# 	pass


	# for i in range((int(total_count)//100)+2):
	# 	# print(i)
	# 	my_api.get_representations(mms_id,{"limit":"100","offset":99*i})
	# 	repres = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)
	# 	# print(repres)
	# 	for rep in repres:
	# 		# print(rep)
	# 		try:
	# 			my_api.get_representation (mms_id, rep)
	# 		except:
	# 			sleep(2)
	# 			my_api.get_representation (mms_id, rep)
	# 		# ie = re.findall(r"pubam:(.*?)</",my_api.xml_response_data)[0]
	# 		# try:
	# 		# 	year = re.findall(r"year>(.*?)</year",my_api.xml_response_data)[0]
	# 		# 	# print(year)
	# 		# except:
	# 		# 	pass

				
	# 		label = re.findall(r"label>(.*?)</label",my_api.xml_response_data)[0]
	# 		if "2022" in label:# or "2021" in label:
						
	# s 			print(label)
				
				# print(ie)
				# if "2022" in my_api.xml_response_data:
				# 	print(label)
				# 	print(my_api.xml_response_data)
		# my_api.get_representation(mms_id,rep)
		# new_data = my_api.xml_response_data.replace("label>2919 08 </label", "label>2022 01</label")
		# my_api.update_representation(mms_id, rep, new_data)
		# print(my_api.xml_response_data)
		# 			# quit()
		# 			labels[label] = ie
		# 		else:
		# 			labels_to_delete[label] = ie
		# for el in labels_to_delete:
		# 	print(labels_to_delete[el])
		# 	print(labels_to_delete[el])
	# for el in os.listdir(r"Y:\ndha\pre-deposit_prod\server_side_deposits\prod\ld_scheduled\oneoff_audio\9919014268602836\content"):
	# 	if "xml" in el:
	# 		my_repres_getter(el.split(".")[0])

			
				
			# if year in["2021"]:
			# 	print(re.findall(r"label>(.*?)</label",my_api.xml_response_data)[0])
			# 	print (ie)
	# with open("report_items_03092021.txt",'r') as f:
	# 	data = f.read()

	# for line in data.split('\n')[:-1]:
	# 	# print(line)
	# 	mms_id = line.split("|")[2]
	# 	holding_id = line.split("|")[3]
	# 	item_id = line.split("|")[4]
	# 	my_api.get_item(mms_id, holding_id, item_id)
	# 	print(my_api.xml_resgponse_data)
	# 	try:
	# 		print(re.findall(r"description>(.*?)</description",my_api.xml_response_data)[0])
	# 	except:
	# 		print(my_api.xml_response_data)

####################################UPDATE REPRESENTATIONS FROM LIST###################################
	# reps =["32366340600002836","32372987750002836","32372987890002836"]
	# mms = ""
	# for rep in reps:
	# 	my_api.get_representation(mms, rep)
	# 	# print(my_api.xml_response_data)
	# 	my_api.xml_response_data
	# 	label = re.findall(r"<label>(.*?)</label>", my_api.xml_response_data)[0]
	# 	if "iss.1" in label:
	# 		new_label = label.replace("iss.1","iss.01").rstrip(" ")
	# 		print(new_label)
	# 		rep_data = my_api.xml_response_data.replace(label,new_label)
	# 		my_api.update_representation(mms",rep, xml_record_data = rep_data)
	# 		print(my_api.status_code)

	# pol = "POL-131442"
	
	# my_api.get_po_line(pol)
	# print(my_api.xml_response_data)
	# vendor_code = "JBDPUB"
	# my_api.get_vendor(vendor_code)
	# print(my_api.xml_response_data)
##############################UPDATING ITEMS####################################################################3
	mms_id = "9919076273402836"
	holding_id = "22363012950002836"
	descriptions = ["2018 01","2018 02","2018 03","2018 04","2018 05","2018 06","2018 07","2018 08","2018 09","2018 10","2018 11","2019 01","2019 02","2019 03","2019 04","2019 06","2019 07","2019 08","2019 09","2019 10","2019 11","2020 05"]
	my_api.get_items(mms_id ,holding_id,{"limit":"100"})
	print(my_api.xml_response_data)
	items = re.findall(r"<pid>(.*?)</pid>", my_api.xml_response_data)
	for el in items:
		my_api.get_item(mms_id,holding_id,el)
		# print(my_api.xml_response_data)
		descr = re.findall(r"<description>(.*?)</description>", my_api.xml_response_data)[0]
		if descr in descriptions:
			print(descr)
			item_pid = re.findall(r"<pid>(.*?)</pid>", my_api.xml_response_data)[0]
			print(item_pid)
	# 		new_descr = " ".join(descr.split(" ")[:-1])
	# 		print(new_descr)
	# 		cron = re.findall(r"<chronology_k>(.*?)</chronology_k>", my_api.xml_response_data)[0]
	# 		new_data = my_api.xml_response_data.replace("k>01</ch","k></ch" ).replace(descr, new_descr)
	# 		print(new_data)
	# 		my_api.update_item("9916487913502836","22218308440002836",el, new_data)
	# 		print(my_api.xml_response_data)
			my_api.delete_item(mms_id, holding_id, item_pid ,{"override":True})
			print(my_api.xml_response_data)

 
	# items =["23379982560002836","23379963540002836","23379963940002836","23379963950002836","23379963570002836","23379966370002836","23379962140002836","23379963470002836","23379982610002836","23379965300002836","23379963890002836","23379984370002836","23379964640002836","23379982660002836","23379964650002836","23379982770002836","23379984320002836","23379963410002836","23379963340002836","23379962090002836","23379963600002836","23379963370002836","23379966450002836","23379965750002836","23379963510002836","23379966470002836","23379963840002836","23379982540002836","23379982650002836","23379966480002836","23379984080002836","23379963830002836","23379983510002836","23379964900002836","23379984440002836","23379961990002836","23379984360002836","23379962000002836","23379962060002836","23379963440002836","23379984270002836","23379963880002836","23379963770002836","23379983540002836","23379963920002836","23379966440002836","23379963480002836","23379963860002836","23379982620002836","23379963490002836","23379984300002836","23379982700002836","23379962030002836","23379982530002836","23379982640002836","23379982670002836","23379963800002836","23379984340002836","23379962020002836","23379962010002836","23379982760002836","23379984450002836","23379963460002836","23379982600002836","23379983530002836","23379963590002836","23379984350002836","23379963420002836","23379965070002836","23379982680002836","23379984390002836","23379984420002836","23379963850002836","23379984410002836","23379962150002836","23379963580002836","23379962130002836","23379963500002836","23379963900002836","23379984430002836","23379962110002836","23379963520002836","23379963430002836","23379982730002836","23379982690002836","23379983520002836","23379964810002836","23379966460002836","23379964660002836","23379963930002836","23379982630002836","23379982550002836","23379984330002836","23379962120002836","23379963530002836","23379962100002836","23379962190002836","23379962160002836","23379984380002836","23379984400002836","23379982590002836","23379963380002836","23379963550002836","23379983590002836","23379982580002836","23379966060002836","23379963560002836","23379982570002836","23379982520002836","23379983550002836","23379963870002836","23379984310002836","23379983600002836","23379983560002836","23379963450002836","23379961960002836"]	

	# mms_id = "9918191969802836"
	# holding_id = "22305544680002836"
	# my_api.get_items(mms_id, holding_id)
	# items = ["23380740110002836"]
	# for item in items:
	# 	my_api.get_item(mms_id, holding_id, item)
	# 	data = my_api.xml_response_data
	# 	my_api.update_item(mms_id, holding_id, item, data, {"generate_description":True})
	# 	print(my_api.status_code)

	# 	print(my_api.xml_response_data)


if __name__ == '__main__':
	main()